
from django.http import HttpResponse
from children.models import CmMaster 
from django.shortcuts import render
import datetime
from django.contrib.auth.decorators import login_required

@login_required
def attendancesheet(request,grade,month,num_weeks):
    firstday=datetime.date(datetime.date.today().year,int(month),1)
    sunday=firstday + datetime.timedelta(days=-firstday.weekday()-1, weeks=1)
    sundays=[]
    for i in range(int(num_weeks)):
        sundays.append("{:%m/%d}".format(sunday))
        sunday=sunday+ datetime.timedelta(weeks=1)
    
    grades=grade.split("+")
    all_season=['Winter','Winter','Spring','Spring','Spring','Summer','Summer','Summer','Fall','Fall','Fall','Winter']
    quarter=all_season[int(month)-1]
    children_list = CmMaster.objects.filter(ssgrade__in=grades,ssactive='Active').order_by('ssgrade','first_last')
    filler=range(len(children_list),15)
    show_grade=len(grades)>1;
    context ={'children_list': children_list, "grade":grade, "sundays":sundays, "quarter":quarter, "show_grade":show_grade, "filler":filler }
    excel=request.GET.get('excel',False);
    return attendancesheet_export_xlsx(request, children_list, grade,sundays, quarter, show_grade) if excel else render(request, 'children/attendancesheet.html', context)

@login_required
def parentemail(request,grade):
    grades=grade.split("+")
    children_list = CmMaster.objects.filter(ssgrade__in=grades,ssactive='Active').exclude(email__isnull=True).exclude(email__exact='').values_list('email', flat=True).distinct()
    context ={'children_list': children_list, "grade":grade, } 
    return render(request, 'children/parentemail.html', context)

@login_required
def parentcontact(request,grade):
    grades=grade.split("+")
    children_list = CmMaster.objects.filter(ssgrade__in=grades,ssactive='Active').order_by('ssgrade','first_last')
    show_grade=len(grades)>1;
    context ={'children_list': children_list, "show_grade":show_grade, "grade":grade, } 
    excel=request.GET.get('excel',False);
    return parentcontact_export_xlsx(request, children_list, grade, show_grade) if excel else render(request, 'children/parentcontact.html', context)

def test(request):
    now = datetime.datetime.now()
    html = "<html><body>It is now %s.</body></html>" % now
    return HttpResponse(html)

@login_required
def attendancesheet_export_xlsx(request, children_list, grade ,sundays, quarter, show_grade):
    import openpyxl
    from openpyxl.cell import get_column_letter
    file_name = "Grade "+grade+"-"+quarter + " Attendance Sheet.xlsx";

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename='+file_name;
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Grade "+grade+" "+quarter

    row_num = 0

    col_num = 0;
    
    c = ws.cell(row=row_num + 1, column=col_num + 1)
    c.value = "Name"
    col_num=col_num+1
    
    if show_grade:
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = "Grade"
        col_num=col_num+1

    c = ws.cell(row=row_num + 1, column=col_num + 1)
    c.value = "Allergy"
    col_num=col_num+1
    
    
    for sunday in sundays:
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = sunday
        c.style.font.bold = True
        col_num=col_num+1

    row_num = row_num + 1
    
    for children in children_list:
        col_num= 0
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = children.first_last
        col_num=col_num+1
    
        if show_grade:
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value =children.ssgrade
            col_num=col_num+1
            
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = children.allergies_medical_conditions_medications

        row_num = row_num + 1

    wb.save(response)
    return response
    
@login_required
def parentcontact_export_xlsx(request, children_list, grade ,show_grade):
    import openpyxl
    from openpyxl.cell import get_column_letter
    file_name = "Grade "+grade+ " Parent Contacts.xlsx";

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename='+file_name;
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Grade "+grade+" Parent Contacts" 

    row_num = 0

    col_num = 0;
    
    c = ws.cell(row=row_num + 1, column=col_num + 1)
    c.value = "Children Name"
    col_num=col_num+1
    
    if show_grade:
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = "Grade"
        col_num=col_num+1
        
    c = ws.cell(row=row_num + 1, column=col_num + 1)
    c.value = "Father Name"
    col_num=col_num+1
    
    c = ws.cell(row=row_num + 1, column=col_num + 1)
    c.value = "Mother Name"
    col_num=col_num+1
    
    c = ws.cell(row=row_num + 1, column=col_num + 1)
    c.value = "Home Phone"
    col_num=col_num+1

    c = ws.cell(row=row_num + 1, column=col_num + 1)
    c.value = "Email"
    col_num=col_num+1

    row_num = row_num + 1
    
    for children in children_list:
        col_num= 0
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = children.first_last
        col_num=col_num+1
    
        if show_grade:
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value =children.ssgrade
            col_num=col_num+1
            
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = children.fathers_english_name
        col_num=col_num+1

        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = children.mothers_english_name
        col_num=col_num+1

        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = children.home
        col_num=col_num+1

        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = children.email
        col_num=col_num+1

        row_num = row_num + 1

    wb.save(response)
    return response
